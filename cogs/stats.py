import os
from discord.ext import commands
from utils.library import base_functions, profile_functions as pl
from discord import Embed, Member, File
from utils import check
from enum import Enum
import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter

config = base_functions.get_config()


class League(Enum):
    Bronze = "Бронза"
    Silver = "Серебро"
    Gold = "Золото"
    Platinum = "Платина"
    Diamond = "Алмаз"
    Master = "Мастер"
    Grandmaster = "Грандмастер"

class Stats(commands.Cog, name="Stats"):
    """
    — Просмотр таблиц лидеров
    """
    def __init__(self, bot):
        self.bot = bot

    @commands.group(name="top")
    async def top(self, ctx):
        if ctx.invoked_subcommand is None:
            pass

    @top.command(name="excel")
    @check.is_samuro_dev()
    async def top_excel(self, ctx):
        headings = ['id', 'guild_id', 'Победы', 'Поражения', 'Очки', 'Батлтег']
        filepath = 'UserStats.xlsx'
        con, cur = pl.get_con_cur()
        guild_id = pl.get_guild_id(ctx)
        select = pl.selects.get('usGuild')
        cur.execute(select, (guild_id, ))
        data = cur.fetchall()
        cur.close()
        # Создание документа, страницы
        wb = openpyxl.Workbook()
        ws = wb.active
        # Spreadsheet row and column indexes start at 1
        # so we use "start = 1" in enumerate so
        # we don't need to add 1 to the indexes.
        for colno, heading in enumerate(headings, start=1):
            c = ws.cell(row=1, column=colno)
            c.font = Font(bold=True)
            c.value = heading
        # This time we use "start = 2" to skip the heading row.
        for rowno, row in enumerate(data, start=2):
            for colno, cell_value in enumerate(row, start=1):
                ws.cell(row=rowno, column=colno).value = cell_value
        # Выравнивание длины колонок
        dim_holder = DimensionHolder(worksheet=ws)
        for col in range(ws.min_column, ws.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)
        ws.column_dimensions = dim_holder
        # сохранение, вывод, удаление файла
        wb.save(filepath)
        await ctx.author.send(file=File(filepath))
        os.remove(filepath)
        await ctx.send("Файл отправлен личным сообщением")

    @top.command(name="mmr")
    async def top_mmr(self, ctx, league_type="Мастер", count=10):
        """
        - Лидеры по ммр
        """
        try:
            league = League(league_type)
        except:
            await ctx.send("Выберите корректную лигу")
        con, cur = pl.get_con_cur()
        guild_id = pl.get_guild_id(ctx)
        select = pl.selects.get("PlayersLeague")
        cur.execute(select, (league.name, count))
        records = cur.fetchall()
        embed = Embed(
            title=f"Таблица лиги {league.value}",
            color=config["info"]
        )
        value = ""
        for i, record in enumerate(records):
            value += f"{i + 1}. {pl.get_discord_mention(record.id)} (mmr: {record.mmr})\n"
        embed.add_field(
            name=f"Топ {count} игроков",
            value=value
        )
        await ctx.send(embed=embed)

    @top.command(name="wins")
    async def top_wins(self, ctx, count=10):
        """
        - Лидеры по числу побед
        """
        con, cur = pl.get_con_cur()
        guild_id = pl.get_guild_id(ctx)
        select = pl.selects.get("usWins")
        cur.execute(select, (guild_id, count))
        records = cur.fetchall()
        embed = Embed(
            title=f"Таблица лидеров",
            color=config["info"]
        )
        value = ""
        for i, record in enumerate(records):
            value += f"{i + 1}. {pl.get_discord_mention(record.id)} — {record.win}\n"
        embed.add_field(
            name=f"Топ {count} игроков по числу побед",
            value=value
        )
        await ctx.send(embed=embed)

    @top.command(name="points")
    async def top_points(self, ctx, count=10):
        """
        - Лидеры по заработанным очкам
        """
        con, cur = pl.get_con_cur()
        guild_id = pl.get_guild_id(ctx)
        select = pl.selects.get("usPoints")
        cur.execute(select, (guild_id, count, ))
        records = cur.fetchall()
        embed = Embed(
            title=f"Таблица лидеров",
            color=config["info"]
        )
        value = ""
        # max_indent = len(max([self.bot.get_user(x.id).name for x in records])) + 1  # получать макс длину имени
        for i, record in enumerate(records):
            value += f"{i+1}. {pl.get_discord_mention(record.id)} — {record.points}\n"
        embed.add_field(
            name=f"Топ {count} игроков по числу баллов",
            value=value
        )
        await ctx.send(embed=embed)

    @top.command(name="remove")
    @commands.check_any(commands.has_role(703884580041785344),  # Создатель
                        commands.has_role(703884637755408466),  # Админ
                        commands.has_role(711230509967212564),  # Старший модер
                        commands.has_role(711230315540250624),  # Модер
                        commands.has_role(946480695218429952),  # Samuro_dev
                        commands.has_role(880865537058545686),  # test
                        )
    async def points_remove(self, ctx, user: Member, count=0):
        con, cur = pl.get_con_cur()
        guild_id =pl.get_guild_id(ctx)
        select = pl.selects.get("usIdGuild")
        cur.execute(select, (user.id, guild_id))
        record = cur.fetchone()
        stats = pl.get_stats(record)
        if stats.points < count:
            await ctx.send(f"Недостаточно баллов\n"
                           f"Баллов у {pl.get_discord_mention(stats.id)}: {stats.points}")
        else:
            stats.points -= count
            update = pl.updates.get("StatsPoints")
            cur.execute(update, (stats.points, stats.id, stats.guild_id))
            pl.commit(con)
            await ctx.send(f"Баллы успешно сняты\n"
                           f"Осталось баллов: {stats.points}")

    @points_remove.error
    async def points_handler(self, ctx, error):
        if isinstance(error, commands.errors.MissingRole):
            await ctx.send("Недостаточно прав для выполнения команды")


def setup(bot):
    bot.add_cog(Stats(bot))